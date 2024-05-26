import requests
from aiogram import types
from openpyxl import load_workbook
from typing import Union, Optional
from data.config import BACKEND_URL
from aiogram.enums import ParseMode
from states.states import UserNotFound
from aiogram.utils.markdown import hlink
from aiogram.fsm.context import FSMContext
from aiogram.types import ReplyKeyboardRemove
from data.config import ERROR_NOTIFY_BOT_TOKEN, ERROR_NOTIFY_CHANNEL_ID


async def send_error_notify_(status_code: int, line: int, filename: str, request_type: str = 'POST') -> None:
    message = ("MBG-Store-Bot:\n\n"
               f"Request {request_type} so'rovda xatolik yuz berdi.\n"
               f"{filename}  {line}-qator\n"
               f"requests.status_code: {status_code}")

    requests.post(
        url=f'https://api.telegram.org/bot{ERROR_NOTIFY_BOT_TOKEN}/sendMessage',
        data={'chat_id': ERROR_NOTIFY_CHANNEL_ID, 'text': message}
    )


async def get_user_lang(message: types.Message, state: FSMContext) -> str:
    data = await state.get_data()
    if data.get('language'):
        return data.get('language')
    response = requests.get(f"{BACKEND_URL}/check/?telegram_id={message.from_user.id}")

    if response.json().get("ok") and response.status_code == 200:
        if response.json().get('user'):
            lang = response.json().get('result').get('language')
            await state.update_data({'language': lang})
            return lang
        else:
            if message.text != "/start":
                await message.answer(
                    text="Foydalanuvchilar uchun botga yangi imkoniyatlar qo'shildi 🎉.\n"
                         "Iltimos /start ni bosing.\n\n"
                         "Для пользователей в бот добавлены новые возможности 🎉.\n"
                         "Нажмите пожалуйста /start.",
                    reply_markup=ReplyKeyboardRemove()
                )
            await state.set_state(UserNotFound.user_id)

            return ''
    else:
        await send_error_notify_(
            status_code=response.status_code,
            line=29, filename='assistants.py', request_type='GET'
        )
        await network_error_message(
            message=message
        )
        return ''


async def network_error_message(
        message: types.Message,
        button: Optional[
            Union[
                types.InlineKeyboardMarkup, types.ReplyKeyboardMarkup, types.ReplyKeyboardRemove
            ]
        ] = None
):
    if button is None:
        await message.answer(
            text="Tarmoqda ulanish xatoligi yuz berdi\n"
                 "Iltimos qaytadan urinib ko'ring.\n\n"
                 "Произошла ошибка подключения к сети\n"
                 "Пожалуйста, попробуйте еще раз."
        )
    else:
        await message.answer(
            text="Tarmoqda ulanish xatoligi yuz berdi\n"
                 "Iltimos qaytadan urinib ko'ring.\n\n"
                 "Произошла ошибка подключения к сети\n"
                 "Пожалуйста, попробуйте еще раз.",
            reply_markup=button
        )


async def send_content(message: types.Message, data, lang):
    for content in data.get('content')[:5]:
        img_list = content.get("images")
        loc_name = content.get('store').get('store_location_name')
        if loc_name is None:
            loc_name = content.get('store').get('brand_name')
        product_url = hlink(
            title={
                'uz': "🛍 Mahsulotni xarid qilish",
                'ru': "🛍 Покупка товара"
            }.get(lang),
            url=f"https://www.mbgstore.uz/products/{content.get('id')}/")

        shop_location = hlink(title='📍  ' + loc_name,
                              url=f"https://maps.google.com/maps?"
                                  f"q={content.get('store').get('latitude')},"
                                  f"{content.get('store').get('longitude')}"
                              )
        if content.get('discount'):
            discount_price = {
                'uz': f"\n🎉 Ushbu mahsulot uchun chegirma mavjud\n"
                      f"💰 Mahsulot narxi: <del>{content.get('price')} UZS</del>   "
                      f"<ins>{content.get('discount_price')} UZS</ins>\n\n",

                'ru': f"\n🎉 На этот товар Действует Скидка\n"
                      f"💰 Цена товара: <del>{content.get('price')} UZS</del>   "
                      f"<ins>{content.get('discount_price')} UZS</ins>\n\n"
            }.get(lang)
        else:
            discount_price = {
                'uz': f"💰 Mahsulot narxi: <ins>{content.get('price')} UZS</ins>\n\n",
                'ru': f"💰 Цена товара: <ins>{content.get('price')} UZS</ins>\n\n"
            }.get(lang)

        text = {
            'uz': f"🏭 Do'kon nomi: {content.get('store').get('brand_name')}\n\n"
                  f"Mahsulot nomi: {content.get('name')}\n"
                  f"Reyting darajasi: {content.get('rating')}\n"
                  f"{discount_price}"
                  f"Mahsulot haqida: {content.get('description')}\n\n"
                  f"{product_url}\n"
                  f"{shop_location}",
            'ru': f"🏭 Название магазина: {content.get('store').get('brand_name')}\n\n"
                  f"Название продукта: {content.get('name')}\n"
                  f"Рейтинговый уровень: {content.get('rating')}\n"
                  f"{discount_price}"
                  f"О продукте: {content.get('description')}\n\n"
                  f"{product_url}\n"
                  f"{shop_location}"
        }
        media = [
            types.InputMediaPhoto(
                media=img_list.pop().get('image'),
                caption=text.get(lang),
                parse_mode=ParseMode.HTML
            )]

        media += list(map(lambda img: types.InputMediaPhoto(media=img.get('image')), img_list))

        await message.answer_media_group(
            media=media,
        )


async def send_products_xlsx(telegram_id):
    files = {
        'file': open(f"data/{telegram_id}_products.xlsx", "rb")
    }
    result = requests.post(
        url=f"{BACKEND_URL}/create/product_xlsx/",
        data={
            'telegram_id': telegram_id,
        },
        files=files
    )
    return result


async def check_file_xlsx(telegram_id):
    wb_object = load_workbook(f'data/{telegram_id}_products.xlsx')
    wb_active = wb_object.active
    No = wb_active.cell(column=1, row=1).value == "№"  # A
    barcode = wb_active.cell(column=2, row=1).value == "Штрих-код"  # B
    name = wb_active.cell(column=3, row=1).value == "Название товара"  # C
    category = wb_active.cell(column=4, row=1).value == "Категория товара "  # D
    available = wb_active.cell(column=5, row=1).value == "Количество"  # E
    entry_price = wb_active.cell(column=6, row=1).value == "Себестоимость"  # F
    price = wb_active.cell(column=7, row=1).value == "Цена"  # G
    discount = wb_active.cell(column=8, row=1).value == "Скидка (в процентах)"  # H
    return No and barcode and name and category and available and entry_price and price and discount
